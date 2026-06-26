"""sauron_file_search.py — Поиск ФИО из файлов через Sauron API.

Модуль не выполняет никаких запросов при импорте.
Точка входа: run_file_search(data, filename, chat_id, bot, progress_msg_id)

Формат отчёта: 4 листа XLSX
  1) «Итог по людям»   — одна строка = один исходный человек
  2) «Родственники»    — одна строка = один вероятный родственник
  3) «Проверка номеров»— одна строка = один проверенный номер
  4) «Ошибки и лимиты» — пропуски, ошибки API, лимиты

Env:
  SAURON_API_KEY             — обязательный
  SFS_MAX_FIO                — макс. ФИО за файл      (умолч. 30)
  SFS_DELAY_SEC              — пауза между запросами  (умолч. 2)
  SFS_MAX_RELATIVES          — макс. родственников/ФИО (умолч. 4)
  SFS_TOP_PHONES             — топ номеров на человека (умолч. 4)
  SFS_FAMILY_SCORE_MIN       — порог родства (умолч. 5)
"""
from __future__ import annotations

import io
import os
import re
import csv
import time
import logging
from dataclasses import dataclass, field
from typing import Optional

logger = logging.getLogger(__name__)

# ── Лимиты ────────────────────────────────────────────────────────────────────
MAX_FIO          = int(os.environ.get("SFS_MAX_FIO",         "30"))
DELAY_SEC        = float(os.environ.get("SFS_DELAY_SEC",     "2"))
MAX_RELATIVES    = int(os.environ.get("SFS_MAX_RELATIVES",   "4"))
TOP_PHONES       = int(os.environ.get("SFS_TOP_PHONES",      "4"))
FAMILY_SCORE_MIN = float(os.environ.get("SFS_FAMILY_SCORE_MIN", "5"))

# ── Опциональные библиотеки ───────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

try:
    import xlrd
    _HAS_XLRD = True
except ImportError:
    _HAS_XLRD = False

try:
    from docx import Document as DocxDoc
    _HAS_DOCX = True
except ImportError:
    _HAS_DOCX = False

try:
    import pdfplumber
    _HAS_PDF = True
except ImportError:
    _HAS_PDF = False


# ═════════════════════════════════════════════════════════════════════════════
# Структуры данных
# ═════════════════════════════════════════════════════════════════════════════

@dataclass
class InputRecord:
    row_num: int
    fio:     str
    dob:     str = ""   # дата рождения
    dod:     str = ""   # дата смерти
    phone:   str = ""   # телефон из файла

@dataclass
class PhoneCheck:
    """Один проверенный номер для листа «Проверка номеров»."""
    owner_fio:    str = ""
    owner_type:   str = ""   # "основной" | "родственник"
    phone_norm:   str = ""
    operator:     str = ""
    in_sauron:    bool = False
    in_maxim:     bool = False
    social_links: str = ""
    fresh_src:    bool = False
    status:       str = ""   # "действующий" | "сомнительный" | "отброшен"
    reject_reason:str = ""

@dataclass
class RelativeRecord:
    """Один вероятный родственник для листа «Родственники»."""
    source_fio:      str   = ""   # исходное ФИО из файла
    source_row:      int   = 0    # номер строки исходного человека (для точного маппинга)
    main_fio:        str   = ""   # найденный основной человек
    fio:             str   = ""   # ФИО родственника
    dob:             str   = ""   # ТОЛЬКО дата в формате DD.MM.YYYY
    snils:           str   = ""   # СНИЛС родственника (как текст)
    variants_fio:    str   = ""   # варианты ФИО / алиасы (отдельно от даты!)
    alt_fio:         str   = ""   # прежняя/девичья фамилия (устарело, используй variants_fio)
    relation:        str   = ""   # тип родства — всегда заполнен человекочитаемо
    address:         str   = ""
    phones:          list[str] = field(default_factory=list)
    phone_note:      str   = ""
    vk:              str   = ""
    ok:              str   = ""
    other_social:    str   = ""
    in_maxim:        bool  = False
    maxim_source:    str   = ""
    evidence:        str   = ""   # все доказательства родства
    confidence:      str   = ""   # высокая/средняя/низкая
    social_evidence: str   = ""   # доказательства из соцсетей
    common_signs:    str   = ""   # общие признаки (телефон, адрес, email…)
    score:           float = 0.0  # числовой скоринг родства
    evidence_source: str   = ""   # источник: Sauron, ВКонтакте, ОК и т.д.
    # VK API поля (заполняются если VK_API_TOKEN настроен)
    vk_profile_url:  str   = ""
    vk_full_name:    str   = ""
    vk_maiden_name:  str   = ""
    vk_city:         str   = ""
    vk_relatives_str:str   = ""
    vk_evidence:     str   = ""
    comment:         str   = ""

@dataclass
class PersonRecord:
    """Один исходный человек для листа «Итог по людям»."""
    row_num:     int  = 0
    source_fio:  str  = ""
    dob:         str  = ""
    dod:         str  = ""
    found:       bool = False
    found_fio:   str  = ""
    addresses:   str  = ""
    phones:      str  = ""
    emails:      str  = ""
    vk:          str  = ""
    ok:          str  = ""
    other_social:str  = ""
    in_maxim:    bool = False
    maxim_source:str  = ""
    sources:     str  = ""
    rel_count:   int  = 0
    error:       str  = ""
    comment:     str  = ""

@dataclass
class ErrorRecord:
    """Одна строка ошибки/лимита для листа «Ошибки и лимиты»."""
    row_num:    int  = 0
    source_fio: str  = ""
    error_type: str  = ""
    detail:     str  = ""


# ═════════════════════════════════════════════════════════════════════════════
# Вспомогательные паттерны
# ═════════════════════════════════════════════════════════════════════════════

_NAME_STOPS = frozenset({
    "Республика","Область","Край","Округ","Район","Город","Москва",
    "Улица","Проспект","Бульвар","Переулок","Площадь","Набережная",
    "Январь","Февраль","Март","Апрель","Май","Июнь","Июль","Август",
    "Сентябрь","Октябрь","Ноябрь","Декабрь",
    "Российская","Украина","Украинская","Белорусская",
})
_GEO_RE = re.compile(r'^[А-ЯЁ][а-яё]{3,}(ская|ский|ское|ской)$')
_FIO3_RE = re.compile(
    r'\b([А-ЯЁІЇЄ][а-яёіїє\'\-]{1,25})\s+'
    r'([А-ЯЁІЇЄ][а-яёіїє\'\-]{1,20})\s+'
    r'([А-ЯЁІЇЄ][а-яёіїє\'\-]{1,20})\b'
)
_FIO2_RE = re.compile(
    r'\b([А-ЯЁІЇЄ][а-яёіїє\'\-]{1,25})\s+'
    r'([А-ЯЁІЇЄ][а-яёіїє\'\-]{1,20})\b'
)
_DATE_RE = re.compile(r'\b(\d{1,2}[./]\d{1,2}[./]\d{2,4}|\d{4})\b')

# Семейные ключевые слова
_FAMILY_KWS = frozenset({
    'супруг','супруга','жена','муж','брат','сестра','дочь','сын',
    'мать','отец','мама','папа','родители','дедушка','бабушка',
    'внук','внучка','племянник','племянница','тётя','дядя',
    'сноха','зять','теща','тёща','тесть','свекровь','свёкор','свекор',
    'родственник','ребёнок','дети','родитель','отпрыск','наследник',
})
_NON_FAMILY_KWS = frozenset({
    'клиент','водитель','коллега','сотрудник','работодатель',
    'партнёр','учредитель','директор','акционер','владелец',
    'знакомый','друг','подруга','осужденный','подозреваемый',
})
_FAMILY_SRC_KWS = (
    'загс','браки','бракосочетание','свидетельство о рождении',
    'семейная','алименты','материнский капитал','опека',
    'пфр','пенсионный','родственник',
)

# Maxim / taxsee
_MAXIM_RE = re.compile(
    r'taxsee|\bmaxim(um)?\b|максим(а|у|ом|е|овск)?\b|'
    r'mac[s]?tax|мак[сш]такс|maxtax|'
    r'клиент[ыа]?\s*(maxim|макс|такси)|база\s*(maxim|макс)|driver\s*base',
    re.IGNORECASE,
)

# Социальные сети
_VK_RE  = re.compile(r'vk\.com/[\w.\-]+|(?:id|club|public)\d{4,}|vkontakte\.ru/[\w.]+', re.I)
_OK_RE  = re.compile(r'ok\.ru/[\w.\-]+|odnoklassniki\.ru/[\w.]+', re.I)
_SOC_RE = re.compile(
    r'(?:instagram|facebook|twitter|telegram|t\.me|youtube|tiktok|linkedin)[./][\w.\-/]+',
    re.I,
)

# Телефоны
_ALL_SAME_RE = re.compile(r'^(\d)\1{9,}$')
_SEQUENTIAL  = {'1234567890','0987654321','9876543210','0123456789'}
_TECH_PFX    = ('70000','79000','700000','71234','70123')
_RU_MOB_CODES = frozenset({
    '901','902','903','904','905','906','908','909',
    '910','911','912','913','914','915','916','917','918','919',
    '920','921','922','923','924','925','926','927','928','929',
    '930','931','932','933','934','936','937','938','939',
    '950','951','952','953','955','958','960','961','962','963',
    '964','965','966','967','968','969',
    '970','977','978','980','981','982','983','984','985','986',
    '987','988','989','990','991','992','993','994','995','996',
    '997','998','999',
})

# Поля-алиасы в ответе Sauron
_ALIAS_FIELDS = (
    'Прежняя фамилия','Другие ФИО','Алиас','Девичья',
    'Старая фамилия','Псевдоним','old_name','alias',
    'Другие имена','ФИО ранее','Предыдущая фамилия',
    'Maiden name','Previous name',
)


# ═════════════════════════════════════════════════════════════════════════════
# Телефонные утилиты
# ═════════════════════════════════════════════════════════════════════════════

def _norm_phone(raw: str) -> Optional[str]:
    if not raw:
        return None
    digits = re.sub(r'\D', '', str(raw))
    if not digits or not (7 <= len(digits) <= 15):
        return None
    if len(digits) == 11 and digits[0] in ('7','8'):
        norm = '7' + digits[1:]
    elif len(digits) == 10 and digits[0] == '0':
        norm = '380' + digits[1:]
    elif len(digits) == 12 and digits.startswith('380'):
        norm = digits
    else:
        norm = digits
    if _ALL_SAME_RE.match(norm):
        return None
    if norm[-10:] in _SEQUENTIAL:
        return None
    if any(norm.startswith(p) for p in _TECH_PFX):
        return None
    if norm.replace('0','') == '':
        return None
    if len(norm) == 11 and norm.startswith('7'):
        if norm[1] not in ('9','4','8','3','2'):
            return None
    return norm


def _is_mobile(norm: str) -> bool:
    if norm and norm.startswith('79') and len(norm) == 11:
        return norm[1:4] in _RU_MOB_CODES
    if norm and norm.startswith('380') and len(norm) == 12:
        return norm[3] in ('6','7','9','5','4','3')
    return False


def _phone_score(raw: str, src_ctx: str, freq: int = 1) -> float:
    norm = _norm_phone(raw)
    if norm is None:
        return -1.0
    score = 10.0 if _is_mobile(norm) else 2.0
    for yr in ('2026','2025','2024'):
        if yr in src_ctx:
            score += 5.0
            break
    for yr in ('2023','2022'):
        if yr in src_ctx:
            score += 2.0
            break
    if _MAXIM_RE.search(src_ctx):
        score += 8.0
    score += min(freq, 5) * 1.5
    return score


def _pick_top_phones(
    phones_raw: list[str],
    freq:  dict[str, int],
    src:   dict[str, list[str]],
    seen:  set[str],
    top_n: int = TOP_PHONES,
) -> tuple[list[str], int, str]:
    """Топ-N ликвидных мобильных номеров по скорингу."""
    scored: list[tuple[float, str]] = []
    seen_local: set[str] = set()
    rejected = 0
    reject_reasons: list[str] = []

    for raw in phones_raw:
        norm = _norm_phone(raw)
        if norm is None:
            rejected += 1
            reject_reasons.append("невалидный")
            continue
        if norm in seen or norm in seen_local:
            rejected += 1
            reject_reasons.append("дубль")
            continue
        if not _is_mobile(norm):
            rejected += 1
            reject_reasons.append("городской")
            continue
        seen_local.add(norm)
        ctx = ' '.join(src.get(raw, []))
        score = _phone_score(raw, ctx, freq.get(raw, 1))
        scored.append((score, norm))

    scored.sort(key=lambda x: -x[0])
    top = scored[:top_n]
    for _, norm in top:
        seen.add(norm)

    liquid = [n for _, n in top]
    discarded = len(phones_raw) - len(liquid)

    notes = []
    if rejected > 0:
        notes.append(f"отброшено {rejected} ({', '.join(dict.fromkeys(reject_reasons[:3]))})")
    if len(scored) > top_n:
        notes.append(f"оставлено топ {top_n}")

    return liquid, discarded, '; '.join(notes)


# ═════════════════════════════════════════════════════════════════════════════
# Утилиты для ФИО
# ═════════════════════════════════════════════════════════════════════════════

def _valid_word(w: str) -> bool:
    if w in _NAME_STOPS or _GEO_RE.match(w):
        return False
    return 2 <= len(w) <= 30


def _extract_fios(text: str) -> list[str]:
    found, seen = [], set()
    for m in _FIO3_RE.finditer(text):
        parts = [m.group(1), m.group(2), m.group(3)]
        if all(_valid_word(p) for p in parts):
            name = ' '.join(parts)
            if name not in seen:
                seen.add(name); found.append(name)
    for m in _FIO2_RE.finditer(text):
        p1, p2 = m.group(1), m.group(2)
        if _valid_word(p1) and _valid_word(p2):
            name = f'{p1} {p2}'
            if name not in seen and not any(name in ex for ex in seen):
                seen.add(name); found.append(name)
    return found


def _fio_last(fio: str) -> str:
    parts = fio.strip().split()
    return parts[0].lower() if parts else ''


def _fio_patron(fio: str) -> str:
    parts = fio.strip().split()
    return parts[2].lower() if len(parts) >= 3 else ''


# ═════════════════════════════════════════════════════════════════════════════
# Парсинг входного файла
# ═════════════════════════════════════════════════════════════════════════════

def _decode(data: bytes) -> str:
    for enc in ('utf-8','cp1251','utf-16','latin-1'):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode('utf-8', errors='replace')


_FIO_COLS    = frozenset({'фио','fullname','full_name','ф.и.о','ф.и.о.','полноеимя','person','человек','фамилияимяотчество'})
_LAST_COLS   = frozenset({'фамилия','lastname','last_name','surname','фам'})
_FIRST_COLS  = frozenset({'имя','firstname','first_name','first'})
_PATRON_COLS = frozenset({'отчество','patronymic','patronym','middlename','middle_name','middle'})
_DOB_COLS    = frozenset({'датарождения','дата_рождения','дата рождения','birth','dob','birthday','датарожд','др'})
_DOD_COLS    = frozenset({'датасмерти','дата_смерти','датасмерт','death','dod','датасм'})
_PHONE_COLS  = frozenset({'телефон','phone','тел','моб','mob','mobile','номер'})


def _norm_hdr(s: str) -> str:
    return re.sub(r'[\s_.\-]+', '', str(s).lower().strip())


def _detect_cols(headers: list[str]) -> dict[str, int]:
    r = {k: -1 for k in ('fio','last','first','patron','dob','dod','phone')}
    for i, h in enumerate(headers):
        n = _norm_hdr(h)
        if n in _FIO_COLS and r['fio'] < 0:
            r['fio'] = i
        elif n in _LAST_COLS and r['last'] < 0:
            r['last'] = i
        elif n in _FIRST_COLS and r['first'] < 0:
            r['first'] = i
        elif n in _PATRON_COLS and r['patron'] < 0:
            r['patron'] = i
        elif n in _DOB_COLS and r['dob'] < 0:
            r['dob'] = i
        elif n in _DOD_COLS and r['dod'] < 0:
            r['dod'] = i
        elif n in _PHONE_COLS and r['phone'] < 0:
            r['phone'] = i
    return r


def _cell(row: list[str], idx: int) -> str:
    return str(row[idx]).strip() if 0 <= idx < len(row) else ''


def _rows_to_records(headers: list[str], rows: list[list[str]]) -> list[InputRecord]:
    col = _detect_cols(headers)
    has_struct = any(v >= 0 for v in (col['fio'], col['last'], col['first']))
    records: list[InputRecord] = []
    seen: set[str] = set()

    def _add(row_num: int, fio: str, dob: str, dod: str, phone: str):
        key = fio.strip().lower()
        if not key or key in seen:
            return
        seen.add(key)
        records.append(InputRecord(row_num=row_num, fio=fio.strip(),
                                   dob=dob, dod=dod, phone=phone))

    for i, row in enumerate(rows, 2):
        dob   = _cell(row, col['dob'])
        dod   = _cell(row, col['dod'])
        phone = _cell(row, col['phone'])

        if has_struct:
            if col['fio'] >= 0:
                fio = _cell(row, col['fio'])
            else:
                fio = ' '.join(p for p in [
                    _cell(row, col['last']),
                    _cell(row, col['first']),
                    _cell(row, col['patron']),
                ] if p)
            if fio:
                _add(i, fio, dob, dod, phone)
        else:
            line = ' '.join(str(c) for c in row if str(c).strip())
            for name in _extract_fios(line):
                _add(i, name, dob, dod, phone)

    return records


def parse_input_file(data: bytes, filename: str) -> tuple[list[InputRecord], str]:
    """Парсит файл и возвращает (records, error_msg)."""
    fname = filename.lower()
    records: list[InputRecord] = []

    try:
        if fname.endswith('.csv'):
            text = _decode(data)
            for sep in (',', ';', '\t', '|'):
                reader = csv.reader(io.StringIO(text), delimiter=sep)
                rows = [r for r in reader if any(c.strip() for c in r)]
                if len(rows) >= 2 and len(rows[0]) >= 2:
                    records = _rows_to_records(rows[0], rows[1:])
                    break
            if not records:
                text_lines = [l.strip() for l in text.splitlines() if l.strip()]
                for i, line in enumerate(text_lines, 1):
                    for name in _extract_fios(line):
                        if name.lower() not in {r.fio.lower() for r in records}:
                            records.append(InputRecord(row_num=i, fio=name))

        elif fname.endswith('.xlsx'):
            if not _HAS_OPENPYXL:
                return [], "xlsx не поддерживается — установи openpyxl"
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            rows = [[str(c.value or '').strip() for c in row] for row in ws.iter_rows()]
            wb.close()
            if len(rows) >= 2:
                records = _rows_to_records(rows[0], rows[1:])
            elif rows:
                for i, row in enumerate(rows, 1):
                    for name in _extract_fios(' '.join(row)):
                        records.append(InputRecord(row_num=i, fio=name))

        elif fname.endswith('.xls'):
            if not _HAS_XLRD:
                return [], "xls не поддерживается — установи xlrd"
            wb = xlrd.open_workbook(file_contents=data)
            ws = wb.sheet_by_index(0)
            rows = [[str(ws.cell_value(r, c)).strip() for c in range(ws.ncols)] for r in range(ws.nrows)]
            if len(rows) >= 2:
                records = _rows_to_records(rows[0], rows[1:])

        elif fname.endswith('.txt'):
            text = _decode(data)
            seen: set[str] = set()
            for i, line in enumerate(text.splitlines(), 1):
                line = line.strip()
                if not line:
                    continue
                # Попытка найти дату рождения в строке
                dob = ''
                dm = _DATE_RE.search(line)
                if dm:
                    dob = dm.group(1)
                for name in _extract_fios(line):
                    key = name.lower()
                    if key not in seen:
                        seen.add(key)
                        records.append(InputRecord(row_num=i, fio=name, dob=dob))

        elif fname.endswith('.docx'):
            if not _HAS_DOCX:
                return [], "docx не поддерживается — установи python-docx"
            doc = DocxDoc(io.BytesIO(data))
            lines = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
            seen2: set[str] = set()
            for i, line in enumerate(lines, 1):
                for name in _extract_fios(line):
                    if name.lower() not in seen2:
                        seen2.add(name.lower())
                        records.append(InputRecord(row_num=i, fio=name))

        elif fname.endswith('.pdf'):
            if not _HAS_PDF:
                return [], "pdf не поддерживается — установи pdfplumber"
            lines = []
            with pdfplumber.open(io.BytesIO(data)) as pdf:
                for page in pdf.pages[:20]:
                    t = page.extract_text()
                    if t:
                        lines.extend(t.splitlines())
            seen3: set[str] = set()
            for i, line in enumerate(lines, 1):
                for name in _extract_fios(line.strip()):
                    if name.lower() not in seen3:
                        seen3.add(name.lower())
                        records.append(InputRecord(row_num=i, fio=name))
        else:
            ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else '?'
            return [], f"Формат .{ext} не поддерживается.\nПоддержка: txt, csv, xlsx, xls, docx, pdf"

    except Exception as e:
        return [], f"Ошибка чтения файла: {str(e)[:120]}"

    if not records:
        return [], (
            "В файле не найдено ни одного ФИО.\n\n"
            "Убедись, что:\n"
            "• Для csv/xlsx/xls — есть колонки «ФИО», «Фамилия», «Имя» или «Отчество»\n"
            "• Для txt/docx/pdf — каждая строка содержит 2–3 слова с заглавной буквы (кириллица)"
        )
    return records, ''


# ═════════════════════════════════════════════════════════════════════════════
# Извлечение полей из ответа Sauron
# ═════════════════════════════════════════════════════════════════════════════

def _val(rec: dict, *keys: str) -> str:
    for k in keys:
        v = rec.get(k)
        if v and str(v).strip() and str(v).strip().lower() not in ('none','null','-','nan'):
            return str(v).strip()
    return ''


def _extract_fields(api_result: dict) -> dict:
    """Извлекает все полезные поля из ответа Sauron."""
    records = api_result.get("response", [])
    if not records:
        return {
            "found_fio":"","phones_raw":[],"phone_freq":{},"phone_src":{},
            "emails":"","addresses":"","connections_parsed":[],"aliases":[],
            "sources":"","raw_count":0,"raw_records":[],
            "vk":"","ok":"","other_social":"",
        }

    all_fios, all_phones, all_emails, all_addrs, all_srcs = [], [], [], [], []
    conn_counts: dict[str, int] = {}
    phone_freq:  dict[str, int] = {}
    phone_src:   dict[str, list[str]] = {}
    alias_set:   set[str] = set()

    for rec in records:
        fio = _val(rec, 'ФИО','Фамилия','Имя')
        if fio and fio not in all_fios:
            all_fios.append(fio)

        src = _val(rec, 'Источник','База','Источники')
        if src and src not in all_srcs:
            all_srcs.append(src)

        for pf in ('Телефон','Телефон2','Телефон3','Phone','Моб','Mob','Тел'):
            ph = _val(rec, pf)
            if ph:
                phone_freq[ph] = phone_freq.get(ph, 0) + 1
                phone_src.setdefault(ph, []).append(src)
                if ph not in all_phones:
                    all_phones.append(ph)

        for ef in ('Email','E-mail','Эл. почта','Почта'):
            em = _val(rec, ef)
            if em and em not in all_emails:
                all_emails.append(em)

        addr_parts = []
        for af in ('Страна','Регион','Город','Населенный пункт','Адрес','Улица','Дом','Квартира'):
            v = _val(rec, af)
            if v:
                addr_parts.append(v)
        if addr_parts:
            addr = ', '.join(addr_parts)
            if addr not in all_addrs:
                all_addrs.append(addr)

        conn = _val(rec, 'Связь с лицом','Связанные лица','Связи','Родственники')
        if conn:
            # Разбиваем многолюдные строки «ФИО1; ФИО2; ФИО3» на отдельные записи
            for part in re.split(r'\s*;\s*', conn):
                part = part.strip()
                if part and len(part) > 3:
                    conn_counts[part] = conn_counts.get(part, 0) + 1

        # Алиасы / предыдущие фамилии
        for af in _ALIAS_FIELDS:
            v = _val(rec, af)
            if v and v.lower() not in ('none','null','-'):
                for nm in _extract_fios(v):
                    alias_set.add(nm)

    def _join(lst: list[str], n: int = 8) -> str:
        return '; '.join(lst[:n])

    # Парсинг связей
    connections_parsed: list[dict] = []
    for raw, cnt in sorted(conn_counts.items(), key=lambda x: -x[1]):
        fio, dob, rel = _parse_conn(raw)
        connections_parsed.append({
            'fio': fio, 'dob': dob, 'relation': rel,
            'raw': raw, 'count': cnt,
        })

    # Соцсети
    vk, ok, other = _extract_social(records)

    return {
        "found_fio":         _join(all_fios, 3),
        "phones_raw":        all_phones,
        "phone_freq":        phone_freq,
        "phone_src":         phone_src,
        "emails":            _join(all_emails, 3),
        "addresses":         _join(all_addrs, 5),
        "connections_parsed":connections_parsed,
        "aliases":           list(alias_set)[:5],
        "sources":           _join(all_srcs, 10),
        "raw_count":         len(records),
        "raw_records":       records,
        "vk":                vk,
        "ok":                ok,
        "other_social":      other,
    }


_CONN_RE = re.compile(
    r'^([А-ЯЁІЇЄа-яёіїє][А-ЯЁІЇЄа-яёіїє\'\-]{1,25}\s+'
    r'[А-ЯЁІЇЄа-яёіїє][А-ЯЁІЇЄа-яёіїє\'\-]{1,20}'
    r'(?:\s+[А-ЯЁІЇЄа-яёіїє][А-ЯЁІЇЄа-яёіїє\'\-]{1,20})?)'
    r'(?:\s*[\(\[](.*?)[\)\]])?'
)


def _to_title(s: str) -> str:
    """МУКАШОВ ВЛАДИМИР → Мукашов Владимир (для regex-совместимости)."""
    if s == s.upper() and re.search(r'[А-ЯЁ]{2,}', s):
        return ' '.join(w.capitalize() for w in s.split())
    return s


def _parse_conn(conn: str) -> tuple[str, str, str]:
    """
    «Иванов ИО (01.01.1980/супруг)» → (fio, dob, relation).
    Обрабатывает ALL-CAPS строки типа «МУКАШОВ ВЛАДИМИР ИВАНОВИЧ».
    Дата возвращается ТОЛЬКО если есть цифровой паттерн DD.MM.YYYY.
    Алиасы/варианты ФИО — НЕ попадают в dob.
    """
    raw = conn.strip()
    normalized = _to_title(raw)

    m = _CONN_RE.match(normalized)
    if not m:
        # Fallback: попробуем вытащить ФИО через _extract_fios
        dm = _DATE_RE.search(raw)
        dob = dm.group(1) if dm else ''
        fios = _extract_fios(normalized) or _extract_fios(raw)
        if fios:
            return fios[0], dob, ''
        # Если нет ФИО — возвращаем title-версию, дата пустая
        return normalized, dob, ''

    fio   = m.group(1).strip()
    extra = (m.group(2) or '').strip()
    dob   = ''
    dm    = _DATE_RE.search(extra)
    if dm:
        dob   = dm.group(1)
        extra = extra[:dm.start()] + extra[dm.end():]
    relation = re.sub(r'[/,;]', ' ', extra).strip()
    return fio, dob, relation


def _extract_social(records: list[dict]) -> tuple[str, str, str]:
    vk_s, ok_s, oth_s = set(), set(), set()
    for rec in records:
        for v in rec.values():
            s = str(v or '')
            if len(s) < 5:
                continue
            for m in _VK_RE.finditer(s):
                vk_s.add(m.group(0)[:80])
            for m in _OK_RE.finditer(s):
                ok_s.add(m.group(0)[:80])
            for m in _SOC_RE.finditer(s):
                oth_s.add(m.group(0)[:80])
    return (
        '; '.join(sorted(vk_s)[:4]),
        '; '.join(sorted(ok_s)[:4]),
        '; '.join(sorted(oth_s)[:4]),
    )


def _check_maxim(sources: str, raw_records: Optional[list[dict]] = None) -> tuple[bool, str]:
    combined = sources or ''
    if raw_records:
        for rec in raw_records:
            combined += '\n' + str(rec.get('Источник','') or rec.get('База','') or '')
    m = _MAXIM_RE.search(combined)
    if m:
        s = max(0, m.start()-15)
        e = min(len(combined), m.end()+15)
        return True, combined[s:e].replace('\n',' ').strip()
    return False, ''


# ═════════════════════════════════════════════════════════════════════════════
# Вспомогательные функции для скоринга
# ═════════════════════════════════════════════════════════════════════════════

_DOB_RAW_KEYS = (
    'Дата рождения','ДР','Рождение','Birth','DOB','birthday',
    'дата_рождения','birth_date','Год рождения','Дата рожд',
)

# ── Определение мобильного оператора по коду (best-effort, без учёта MNP) ─────
_RU_OPERATOR: dict[str, str] = {}
for _c in "910 911 912 913 914 915 916 917 918 919 980 981 982 983 984 985 986 987 988 989 891 892 978".split():
    _RU_OPERATOR[_c] = "МТС"
for _c in "903 905 906 909 960 961 962 963 964 965 966 967 968".split():
    _RU_OPERATOR[_c] = "Билайн"
for _c in "920 921 922 923 924 925 926 927 928 929 930 931 932 933 934 935 936 937 938 939 999".split():
    _RU_OPERATOR[_c] = "МегаФон"
for _c in "900 901 902 904 908 950 951 952 953 958 977 991 992 993 994 995 996".split():
    _RU_OPERATOR[_c] = "Tele2"

_UA_OPERATOR: dict[str, str] = {}
for _c in "67 68 96 97 98".split():
    _UA_OPERATOR[_c] = "Київстар"
for _c in "50 66 95 99".split():
    _UA_OPERATOR[_c] = "Vodafone"
for _c in "63 73 93".split():
    _UA_OPERATOR[_c] = "lifecell"


def _phone_operator(norm: str) -> str:
    """Возвращает название оператора по нормализованному номеру или ''."""
    if not norm:
        return ""
    if norm.startswith('7') and len(norm) == 11:
        return _RU_OPERATOR.get(norm[1:4], "")
    if norm.startswith('380') and len(norm) == 12:
        return _UA_OPERATOR.get(norm[3:5], "")
    return ""


_SNILS_KEYS = (
    'СНИЛС', 'Снилс', 'снилс', 'SNILS', 'Snils', 'snils',
    'Номер СНИЛС', 'Страховой номер', 'Страховой номер счёта',
)
_SNILS_RE = re.compile(r'(\d{3})[\-\s]?(\d{3})[\-\s]?(\d{3})[\-\s]?(\d{2})')


def _extract_snils_from_raw(raw_records: list[dict]) -> str:
    """Ищет СНИЛС в ответе Sauron. Возвращает 'XXX-XXX-XXX YY' или ''."""
    for rec in raw_records:
        v = _val(rec, *_SNILS_KEYS)
        if not v:
            continue
        m = _SNILS_RE.search(str(v))
        if m:
            return f"{m.group(1)}-{m.group(2)}-{m.group(3)} {m.group(4)}"
        digits = re.sub(r'\D', '', str(v))
        if len(digits) == 11:
            return f"{digits[0:3]}-{digits[3:6]}-{digits[6:9]} {digits[9:11]}"
    return ''


def _extract_dob_from_raw(raw_records: list[dict]) -> str:
    """Ищет дату рождения в полях ответа Sauron. Возвращает DD.MM.YYYY или ''."""
    for rec in raw_records:
        for k in _DOB_RAW_KEYS:
            v = _val(rec, k)
            if v:
                dm = _DATE_RE.search(str(v))
                if dm:
                    raw_d = dm.group(1)
                    parts = re.split(r'[./]', raw_d)
                    if len(parts) == 3:
                        d, mo, y = parts
                        if len(y) == 2:
                            y = ('19' if int(y) > 30 else '20') + y
                        return f"{d.zfill(2)}.{mo.zfill(2)}.{y}"
                    return raw_d
    return ''


def _split_field(s: str) -> set[str]:
    """Разбивает поле по ';' в set нижнего регистра для сравнения."""
    if not s:
        return set()
    return {v.strip().lower() for v in s.split(';') if v.strip()}


def _human_relation(rel_str: str, evidence: str) -> str:
    """
    Строит читаемое объяснение типа родства.
    Если rel_str пустой или выглядит как ФИО — берём из evidence.
    """
    if rel_str and rel_str not in ('alias/прежнее ФИО',):
        r_low = rel_str.lower()
        if any(kw in r_low for kw in _FAMILY_KWS):
            return rel_str
        # Если строка выглядит как ФИО (нормальный или ALL-CAPS регистр) — игнорируем как тип
        _norm = _to_title(rel_str)  # нормализуем ALL-CAPS для regex
        _looks_fio = (
            len(rel_str) > 15
            and not re.search(r'\d', rel_str)
            and bool(re.search(r'[А-ЯЁ][а-яё]{2,}\s+[А-ЯЁ][а-яё]{1,}', _norm))
        )
        if not _looks_fio and rel_str:
            return rel_str
        # иначе — падаем ниже, строим из evidence

    parts: list[str] = []
    m = re.search(r'явная связь «([^»]+)»', evidence)
    if m:
        parts.append(m.group(1))
    if 'общая фамилия' in evidence:
        parts.append('общая фамилия')
    if 'совпадает отчество' in evidence:
        parts.append('совпадение отчества')
    if 'общий адрес регистрации' in evidence:
        parts.append('общий адрес регистрации')
    if 'общий город' in evidence:
        m2 = re.search(r'общий город/регион:\s*([^;]+)', evidence)
        parts.append(f"общий регион ({m2.group(1).strip()})" if m2 else 'общий регион')
    if 'семейный источник' in evidence:
        m3 = re.search(r'семейный источник:\s*(\S+)', evidence)
        parts.append(f"семейный источник ({m3.group(1)})" if m3 else 'семейный источник')
    if 'общий телефон' in evidence:
        parts.append('общий телефон')
    if 'общий профиль ВКонтакте' in evidence:
        parts.append('общий ВКонтакте')
    if 'общий профиль Одноклассники' in evidence:
        parts.append('общие Одноклассники')
    if 'общий email' in evidence:
        parts.append('общий email')
    if rel_str == 'alias/прежнее ФИО' or 'девичья' in evidence.lower() or 'alias' in rel_str.lower():
        parts.append('смена фамилии / алиас')
    if 'VK' in evidence:
        parts.append('подтверждено ВКонтакте')

    return '; '.join(parts) if parts else (
        'смена фамилии / алиас' if rel_str == 'alias/прежнее ФИО'
        else 'связь по данным Sauron'
    )


# ═════════════════════════════════════════════════════════════════════════════
# Расширенный скоринг родства
# ═════════════════════════════════════════════════════════════════════════════

def _gather_evidence(
    relation:      str,
    main_fio:      str,
    rel_fio:       str,
    shared_addr:   bool,
    rel_sources:   str,
    shared_phones: bool,
    confirm_count: int,
    main_fields:   Optional[dict] = None,
    rel_fields:    Optional[dict] = None,
) -> tuple[float, str, str, str, str, str]:
    """
    Расширенный анализ родства — все доступные поля Sauron + соцсети.

    Returns:
        (score, evidence, confidence, social_evidence, common_signs, evidence_sources)

    score >= FAMILY_SCORE_MIN → вероятный родственник.
    Высокая уверенность: score >= 18.
    """
    score        = 0.0
    evidence:    list[str] = []
    social_ev:   list[str] = []
    common_signs:list[str] = []
    ev_sources:  list[str] = []
    rel_low      = relation.lower()

    # ── 1. Явное родственное ключевое слово ──────────────────────────────
    for kw in _FAMILY_KWS:
        if kw in rel_low:
            score += 15.0
            evidence.append(f"явная связь «{relation}»")
            ev_sources.append('Sauron: поле связи')
            break

    # ── 2. Нейтральный тип связи без семейного кw ────────────────────────
    if relation and not any(kw in rel_low for kw in _FAMILY_KWS | _NON_FAMILY_KWS):
        score += 3.0
        evidence.append(f"поле связи: «{relation}»")

    # ── 3. Штраф за неродственную связь ──────────────────────────────────
    for kw in _NON_FAMILY_KWS:
        if kw in rel_low:
            score -= 12.0
            evidence.append(f"неродственная связь: {kw}")
            break

    # ── 4. Общая фамилия ──────────────────────────────────────────────────
    ml, rl = _fio_last(main_fio), _fio_last(rel_fio)
    if ml and rl and ml == rl:
        score += 8.0
        evidence.append("общая фамилия")
        common_signs.append(f"фамилия: {ml}")

    # ── 5. Общее отчество ────────────────────────────────────────────────
    mp, rp = _fio_patron(main_fio), _fio_patron(rel_fio)
    if mp and rp and mp == rp and mp != ml:
        score += 5.0
        evidence.append("совпадает отчество")
        common_signs.append(f"отчество: {mp}")

    # ── 6. Общий адрес регистрации ────────────────────────────────────────
    if shared_addr:
        score += 5.0
        evidence.append("общий адрес регистрации")
        common_signs.append("общий адрес")
        ev_sources.append('Sauron: адрес')

    # ── 7. Семейный источник ──────────────────────────────────────────────
    src_low = rel_sources.lower()
    for kw in _FAMILY_SRC_KWS:
        if kw in src_low:
            score += 6.0
            evidence.append(f"семейный источник: {kw}")
            ev_sources.append(f'Sauron: {kw}')
            break

    # ── 8. Общий телефон ─────────────────────────────────────────────────
    if shared_phones:
        score += 4.0
        evidence.append("общий телефон")
        common_signs.append("общий телефон")
        ev_sources.append('Sauron: телефон')

    # ── 9. Количество подтверждений ───────────────────────────────────────
    if confirm_count >= 2:
        score += 3.0
        evidence.append(f"{confirm_count}× подтверждений в базе")

    # ── Расширенный анализ из полей Sauron ──────────────────────────────
    if main_fields and rel_fields:
        # 10. Общий ВКонтакте
        sh_vk = _split_field(main_fields.get('vk','')) & _split_field(rel_fields.get('vk',''))
        if sh_vk:
            score += 4.0
            lnk = next(iter(sh_vk))
            evidence.append('общий профиль ВКонтакте')
            social_ev.append(f"ВК: {lnk}")
            ev_sources.append('ВКонтакте')

        # 11. Общие Одноклассники
        sh_ok = _split_field(main_fields.get('ok','')) & _split_field(rel_fields.get('ok',''))
        if sh_ok:
            score += 3.0
            lnk = next(iter(sh_ok))
            evidence.append('общий профиль Одноклассники')
            social_ev.append(f"ОК: {lnk}")
            ev_sources.append('Одноклассники')

        # 12. Общие другие соцсети
        sh_soc = (_split_field(main_fields.get('other_social',''))
                  & _split_field(rel_fields.get('other_social','')))
        if sh_soc:
            score += 3.0
            lnk = next(iter(sh_soc))
            evidence.append('общий профиль в соцсетях')
            social_ev.append(f"соцсеть: {lnk}")
            ev_sources.append('соцсеть')

        # 13. Общий email
        sh_em = _split_field(main_fields.get('emails','')) & _split_field(rel_fields.get('emails',''))
        if sh_em:
            score += 5.0
            em = next(iter(sh_em))
            evidence.append('общий email')
            common_signs.append(f"email: {em}")
            ev_sources.append('email')

        # 14. Фамилия основного упоминается в источниках родственника
        if ml and ml in (rel_fields.get('sources','') or '').lower():
            score += 2.0
            evidence.append('фамилия основного в источниках')
            ev_sources.append('Sauron: источники')

        # 15. Общий город/регион (из адресов)
        main_addr_low = (main_fields.get('addresses','') or '').lower()
        rel_addr_low  = (rel_fields.get('addresses','') or '').lower()
        main_cities   = {w for w in re.findall(r'[а-яё]{4,}', main_addr_low)}
        rel_cities    = {w for w in re.findall(r'[а-яё]{4,}', rel_addr_low)}
        _GEO_STOP     = {'улица','город','район','область','проспект','переулок',
                         'квартира','корпус','строение','номер','дома'}
        shared_cities = main_cities & rel_cities - _GEO_STOP
        if shared_cities and not shared_addr:
            city_ex = sorted(shared_cities)[:2]
            score += 2.0
            evidence.append(f"общий город/регион: {', '.join(city_ex)}")
            common_signs.append(f"регион: {', '.join(city_ex)}")

    confidence = ("высокая" if score >= 18 else
                  "средняя"  if score >= FAMILY_SCORE_MIN else "низкая")

    return (
        score,
        '; '.join(evidence),
        confidence,
        '; '.join(social_ev),
        '; '.join(common_signs),
        '; '.join(dict.fromkeys(ev_sources)),
    )


# ═════════════════════════════════════════════════════════════════════════════
# Пакетный поиск
# ═════════════════════════════════════════════════════════════════════════════

def run_file_search(
    data:            bytes,
    filename:        str,
    chat_id:         int,
    bot,
    progress_msg_id: int,
    max_fio:         int = MAX_FIO,
    delay_sec:       float = DELAY_SEC,
) -> tuple[list[PersonRecord], list[RelativeRecord], list[PhoneCheck], list[ErrorRecord], str]:
    """
    Основная функция поиска.
    Возвращает (persons, relatives, phone_checks, errors, stop_reason).
    """
    import sauron as _sauron

    # ── Парсинг файла ─────────────────────────────────────────────────────
    inp_records, parse_err = parse_input_file(data, filename)
    if parse_err:
        return [], [], [], [ErrorRecord(0, filename, "parse_error", parse_err)], parse_err

    to_search = inp_records[:max_fio]
    skipped   = len(inp_records) - len(to_search)
    total     = len(to_search)

    persons:      list[PersonRecord]   = []
    relatives:    list[RelativeRecord] = []
    phone_checks: list[PhoneCheck]     = []
    errors:       list[ErrorRecord]    = []
    stop_reason   = ""

    global_seen: set[str] = set()   # нормализованные телефоны (dedup)
    query_cache: dict[str, dict]  = {}  # кеш запросов

    # ── Вспомогательные функции ───────────────────────────────────────────
    def _do_search(query: str) -> tuple[bool, dict, bool]:
        q = query.strip()
        if not q:
            return False, {}, False
        if q in query_cache:
            f = query_cache[q]
            return f.get('raw_count',0) > 0, f, False
        try:
            api_result = _sauron._api_post_search(q)
            fields     = _extract_fields(api_result)
            stop       = False
            try:
                if float(api_result.get("balance","999")) < 1.0:
                    stop = True
            except Exception:
                pass
            query_cache[q] = fields
            return fields["raw_count"] > 0, fields, stop
        except RuntimeError as e:
            stop = any(kw in str(e) for kw in ("баланс","Неверный","API-ключ","Secrets"))
            return False, {}, stop
        except Exception as e:
            logger.warning(f"sauron search error for {q!r}: {e}")
            return False, {}, False

    def _add_phone_check(norm: str, owner_fio: str, owner_type: str,
                         in_maxim: bool, sources: str, status: str, reason: str = ''):
        fresh = any(yr in sources for yr in ('2024','2025','2026'))
        pc = PhoneCheck(
            owner_fio=owner_fio, owner_type=owner_type,
            phone_norm=norm,
            in_maxim=in_maxim,
            fresh_src=fresh,
            status=status,
            reject_reason=reason,
            in_sauron=True,
        )
        phone_checks.append(pc)

    def _update_progress(i: int, fio: str):
        try:
            pct = (i - 1) * 100 // total
            bar = '▓' * ((i-1)*10//total) + '░' * (10 - (i-1)*10//total)
            skip_note = f"\n_(пропущено по лимиту: {skipped})_" if skipped and i == 1 else ""
            bot.edit_message_text(
                f"🔍 *Поиск по файлу…*\n\n"
                f"Обработано: *{i-1}* / {total}{skip_note}\n"
                f"{bar} {pct}%\n\n"
                f"👤 Ищу: `{fio}`",
                chat_id, progress_msg_id, parse_mode="Markdown",
            )
        except Exception:
            pass

    # ── Основной цикл по исходным ФИО ─────────────────────────────────────
    for idx, inp in enumerate(to_search, 1):
        fio = inp.fio
        _update_progress(idx, fio)

        pr = PersonRecord(row_num=inp.row_num, source_fio=fio, dob=inp.dob, dod=inp.dod)

        # Строим запрос: ФИО + дата рождения если есть
        query = fio
        if inp.dob:
            query = f"{fio} {inp.dob}"

        found, fields, stop = _do_search(query)

        # Fallback без даты рождения
        if not found and inp.dob:
            found, fields, stop = _do_search(fio)

        if stop:
            pr.error = "Остановлено — недостаточно баланса или ошибка API"
            errors.append(ErrorRecord(inp.row_num, fio, "api_stop", pr.error))
            persons.append(pr)
            stop_reason = pr.error
            break

        if not found:
            pr.comment = "Не найден в Sauron"
            persons.append(pr)
            if idx < total:
                time.sleep(delay_sec)
            continue

        # ── Заполняем основной PersonRecord ──────────────────────────────
        pr.found      = True
        pr.found_fio  = fields['found_fio']
        pr.addresses  = fields['addresses']
        pr.emails     = fields['emails']
        pr.sources    = fields['sources']
        pr.vk         = fields['vk']
        pr.ok         = fields['ok']
        pr.other_social = fields['other_social']

        # Телефоны основного
        liquid, disc, note = _pick_top_phones(
            fields['phones_raw'], fields.get('phone_freq',{}),
            fields.get('phone_src',{}), global_seen, top_n=TOP_PHONES,
        )
        pr.phones   = '; '.join(liquid)
        pr.comment  = note if note else ''
        in_mx, mx_src = _check_maxim(fields['sources'], fields.get('raw_records'))
        pr.in_maxim     = in_mx
        pr.maxim_source = mx_src

        for norm in liquid:
            _add_phone_check(norm, fio, "основной", in_mx, fields['sources'], "действующий")

        # ── Родственники ──────────────────────────────────────────────────
        connections = fields.get('connections_parsed', [])
        aliases     = fields.get('aliases', [])
        main_addrs  = set(a.strip() for a in fields['addresses'].split(';') if a.strip())
        main_phones_norm = set(global_seen)  # уже добавлены выше

        # Формируем кандидатов: (fio_query, dob, relation, count, alt_fio)
        candidates: list[tuple[str, str, str, int, str]] = []
        seen_cand: set[str] = set()

        for cp in connections:
            cand_fio = cp['fio']
            if not cand_fio or len(cand_fio.split()) < 2:
                continue
            key = cand_fio.lower()
            if key in seen_cand or key == fio.lower():
                continue
            seen_cand.add(key)
            candidates.append((cand_fio, cp['dob'], cp['relation'], cp['count'], ''))

        # Алиасы → отдельные поиски
        for alias in aliases:
            key = alias.lower()
            if key not in seen_cand and key != fio.lower():
                seen_cand.add(key)
                candidates.append((alias, '', 'alias/прежнее ФИО', 1, alias))

        # Лимит кандидатов
        candidates = candidates[:MAX_RELATIVES * 2]

        rel_count = 0
        for cand_fio, cand_dob, rel_str, confirm_cnt, alt_fio in candidates:
            if rel_count >= MAX_RELATIVES:
                break

            if delay_sec > 0:
                time.sleep(delay_sec)

            rel_found, rel_fields, rel_stop = _do_search(
                f"{cand_fio} {cand_dob}" if cand_dob else cand_fio
            )
            if not rel_found and cand_dob:
                rel_found, rel_fields, rel_stop = _do_search(cand_fio)

            if rel_stop:
                stop_reason = "Остановлено при поиске родственников — баланс или ошибка API"
                errors.append(ErrorRecord(inp.row_num, fio, "api_stop_rel", stop_reason))
                break

            # Общий адрес?
            rel_addrs = set(a.strip() for a in rel_fields.get('addresses','').split(';') if a.strip())
            shared_addr = bool(main_addrs & rel_addrs)

            # Общий телефон?
            rel_norms: set[str] = set()
            for rp in rel_fields.get('phones_raw',[]):
                n = _norm_phone(rp)
                if n:
                    rel_norms.add(n)
            shared_phones = bool(main_phones_norm & rel_norms)

            # ── Расширенный скоринг родства ───────────────────────────────
            score, evidence, confidence, social_ev, common_signs, ev_src = _gather_evidence(
                rel_str, fio, cand_fio,
                shared_addr, rel_fields.get('sources', ''),
                shared_phones, confirm_cnt,
                main_fields=fields,
                rel_fields=rel_fields,
            )

            if score < FAMILY_SCORE_MIN:
                continue

            # ── Телефоны родственника ──────────────────────────────────────
            rel_local_seen: set[str] = set()
            rel_liquid, _, rel_note = _pick_top_phones(
                rel_fields.get('phones_raw', []),
                rel_fields.get('phone_freq', {}),
                rel_fields.get('phone_src', {}),
                rel_local_seen,
                top_n=TOP_PHONES,
            )
            for norm in rel_liquid:
                global_seen.add(norm)

            rel_in_mx, rel_mx_src = _check_maxim(
                rel_fields.get('sources', ''), rel_fields.get('raw_records')
            )
            rel_vk    = rel_fields.get('vk', '')
            rel_ok    = rel_fields.get('ok', '')
            rel_other = rel_fields.get('other_social', '')

            # ── VK API обогащение (только если токен настроен) ────────────
            vk_data: dict = {}
            if rel_vk:
                try:
                    import vk_api_client as _vk
                    if _vk.is_available():
                        vk_data = _vk.enrich_relative(rel_vk, fio, cand_fio, cand_dob) or {}
                        if vk_data.get('vk_score_bonus', 0) > 0:
                            score    += vk_data['vk_score_bonus']
                            evidence += ('; VK: ' + vk_data['vk_evidence']) if vk_data.get('vk_evidence') else ''
                            if score >= 18:
                                confidence = 'высокая'
                            ev_src = (ev_src + '; VK API').strip('; ')
                except Exception as _ve:
                    logger.debug(f"VK enrich skipped: {_ve}")

            # ── Дата рождения — только настоящая дата ─────────────────────
            real_dob   = cand_dob or _extract_dob_from_raw(rel_fields.get('raw_records', []))
            real_snils = _extract_snils_from_raw(rel_fields.get('raw_records', []))

            rr = RelativeRecord(
                source_fio      = fio,
                source_row      = inp.row_num,
                main_fio        = fields['found_fio'],
                fio             = cand_fio,
                dob             = real_dob,                          # ТОЛЬКО дата, никаких ФИО
                snils           = real_snils,
                variants_fio    = alt_fio,
                alt_fio         = alt_fio,
                relation        = _human_relation(rel_str, evidence),  # читаемая причина
                address         = rel_fields.get('addresses', '') if rel_found else '',
                phones          = rel_liquid,
                phone_note      = rel_note or ('мобильные номера не найдены' if not rel_liquid else ''),
                vk              = rel_vk,
                ok              = rel_ok,
                other_social    = rel_other,
                in_maxim        = rel_in_mx,
                maxim_source    = rel_mx_src,
                evidence        = evidence,
                confidence      = confidence,
                social_evidence = social_ev,
                common_signs    = common_signs,
                score           = round(score, 1),
                evidence_source = ev_src,
                vk_profile_url  = vk_data.get('vk_profile_url', ''),
                vk_full_name    = vk_data.get('vk_full_name', ''),
                vk_maiden_name  = vk_data.get('vk_maiden_name', ''),
                vk_city         = vk_data.get('vk_city', ''),
                vk_relatives_str= vk_data.get('vk_relatives', ''),
                vk_evidence     = vk_data.get('vk_evidence', ''),
                comment         = (
                    f"Найдено {rel_fields.get('raw_count', 0)} записей Sauron"
                    if rel_found else "Не найден в Sauron"
                ),
            )
            relatives.append(rr)
            rel_count += 1

            for norm in rel_liquid:
                _add_phone_check(norm, cand_fio, "родственник", rel_in_mx,
                                 rel_fields.get('sources', ''), "действующий")

        pr.rel_count = rel_count

        if stop_reason:
            persons.append(pr)
            break

        persons.append(pr)

        if idx < total:
            time.sleep(delay_sec)

    if skipped > 0:
        errors.append(ErrorRecord(
            0, filename, "limit",
            f"Пропущено по лимиту ({max_fio} ФИО за файл): {skipped} записей"
        ))

    return persons, relatives, phone_checks, errors, stop_reason


# ═════════════════════════════════════════════════════════════════════════════
# Краткая сводка для чата
# ═════════════════════════════════════════════════════════════════════════════

def build_chat_summary(
    persons:   list[PersonRecord],
    relatives: list[RelativeRecord],
    phones:    list[PhoneCheck],
    errors:    list[ErrorRecord],
    stop_reason: str = "",
) -> str:
    total     = len(persons)
    found     = sum(1 for p in persons if p.found)
    not_found = total - found

    rel_total  = len(relatives)
    rel_alias  = sum(1 for r in relatives if r.alt_fio)
    rel_high   = sum(1 for r in relatives if r.confidence == "высокая")
    phones_act = len([p for p in phones if p.status == "действующий"])
    maxim_cnt  = sum(1 for p in persons if p.in_maxim) + sum(1 for r in relatives if r.in_maxim)
    vk_cnt     = sum(1 for p in persons if p.vk) + sum(1 for r in relatives if r.vk)
    ok_cnt     = sum(1 for p in persons if p.ok) + sum(1 for r in relatives if r.ok)
    lim_errors = sum(1 for e in errors if e.error_type == "limit")
    api_errors = sum(1 for e in errors if "stop" in e.error_type or "api" in e.error_type)

    # VK API статус
    vk_enriched = sum(1 for r in relatives if r.vk_profile_url)
    try:
        import vk_api_client as _vk
        vk_status = "включён ✅" if _vk.is_available() else "не настроен"
    except Exception:
        vk_status = "не настроен"

    lines = ["🔍 *Поиск по файлу завершён*\n"]
    lines.append(f"👤 Обработано ФИО: *{total}*")
    lines.append(f"✅ Найдено: *{found}*   |   ❌ Не найдено: *{not_found}*")
    lines.append(f"\n👨‍👩‍👧 Вероятных родственников: *{rel_total}*")
    if rel_high:
        lines.append(f"  ↳ Высокая уверенность: *{rel_high}*")
    if rel_alias:
        lines.append(f"  ↳ Со сменой фамилии/алиасом: *{rel_alias}*")
    lines.append(f"\n📱 Действующих номеров: *{phones_act}*")
    lines.append(f"🚕 В Максе/taxsee: *{maxim_cnt}*")
    lines.append(f"🔵 ВКонтакте: *{vk_cnt}*   |   🟠 ОК: *{ok_cnt}*")
    lines.append(f"🔷 VK API: {vk_status}"
                 + (f" — обогащено: *{vk_enriched}*" if vk_enriched else ""))

    if lim_errors:
        lines.append(f"\n⚠️ Пропущено по лимиту: *{lim_errors} записей*")
    if api_errors:
        lines.append(f"⛔ Ошибок API: *{api_errors}*")
    if stop_reason:
        lines.append(f"\n⛔ _{stop_reason}_")

    # Топ-5 найденных
    top = [p for p in persons if p.found and (p.phones or p.rel_count)]
    if top:
        lines.append("\n*Топ найденных:*")
        for p in top[:5]:
            ph = p.phones.split(';')[0].strip() if p.phones else ''
            mx = " 🚕" if p.in_maxim else ""
            lines.append(
                f"\n👤 `{p.source_fio}`{mx}"
                + (f"\n  📱 {ph}" if ph else "")
                + (f"\n  👨‍👩‍👧 родственников: {p.rel_count}" if p.rel_count else "")
            )
        if len(top) > 5:
            lines.append(f"_…ещё {len(top)-5}_")

    return '\n'.join(lines)


# ═════════════════════════════════════════════════════════════════════════════
# Построение XLSX-отчёта (4 листа)
# ═════════════════════════════════════════════════════════════════════════════

def _make_border():
    s = Side(style='thin', color='BBBBBB')
    return Border(left=s, right=s, top=s, bottom=s)


def _write_header(ws, cols: list[tuple[str, int]], fill, font, wrap):
    for ci, (name, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.font = font; cell.fill = fill
        cell.alignment = wrap; cell.border = _make_border()
        ws.column_dimensions[cell.column_letter].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 38


def _write_row(ws, row_idx: int, vals: list, fill, wrap):
    brd = _make_border()
    for ci, v in enumerate(vals, 1):
        cell = ws.cell(row=row_idx, column=ci, value=v)
        cell.alignment = wrap; cell.border = brd
        if fill:
            cell.fill = fill


def build_xlsx_report(
    persons:   list[PersonRecord],
    relatives: list[RelativeRecord],
    phones:    list[PhoneCheck],
    errors:    list[ErrorRecord],
) -> Optional[bytes]:
    if not _HAS_OPENPYXL:
        return None

    wb = openpyxl.Workbook()

    # ── Стили ────────────────────────────────────────────────────────────
    H1   = PatternFill("solid", fgColor="1F4E79")   # тёмно-синий — лист 1
    H2   = PatternFill("solid", fgColor="375623")   # тёмно-зелёный — лист 2
    H3   = PatternFill("solid", fgColor="7B2D8B")   # фиолетовый — лист 3
    H4   = PatternFill("solid", fgColor="843C0C")   # коричневый — лист 4
    HFNT = Font(bold=True, color="FFFFFF")

    FOUND    = PatternFill("solid", fgColor="E2EFDA")
    NOTFND   = PatternFill("solid", fgColor="FCE4D6")
    MAXIM_F  = PatternFill("solid", fgColor="FFF2CC")
    REL_BG   = PatternFill("solid", fgColor="EEF5FB")
    REL_MX   = PatternFill("solid", fgColor="FFF0A0")
    REL_HIGH = PatternFill("solid", fgColor="D9F2E6")
    PH_ACT   = PatternFill("solid", fgColor="E2EFDA")
    PH_DIS   = PatternFill("solid", fgColor="FCE4D6")
    ERR_F    = PatternFill("solid", fgColor="FCE4D6")
    WRAP = Alignment(wrap_text=True, vertical="top")

    # ══════════════════════════════════════════════════════════════════════
    # Лист 1: «Итог по людям»
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Итог по людям"
    cols1 = [
        ("Строка №",          7),
        ("Исходное ФИО",     28),
        ("Дата рождения",    14),
        ("Дата смерти",      12),
        ("Найдено",           8),
        ("Найденное ФИО",   28),
        ("Адреса",           38),
        ("Лучшие телефоны", 32),
        ("Email",            20),
        ("ВКонтакте",        22),
        ("Одноклассники",    22),
        ("Другие соцсети",   22),
        ("Есть в Максе",     10),
        ("Источник Макса",   25),
        ("Источники",        30),
        ("Родственников",     9),
        ("Ошибка/Комментарий",28),
    ]
    _write_header(ws1, cols1, H1, HFNT, WRAP)

    for ri, p in enumerate(persons, 2):
        in_mx = p.in_maxim
        fill  = MAXIM_F if in_mx and p.found else (FOUND if p.found else NOTFND)
        _write_row(ws1, ri, [
            p.row_num, p.source_fio, p.dob, p.dod,
            "Да" if p.found else "Нет",
            p.found_fio, p.addresses, p.phones, p.emails,
            p.vk, p.ok, p.other_social,
            "✅ Да" if in_mx else "Нет", p.maxim_source,
            (p.sources or '')[:250],
            p.rel_count,
            (p.error or p.comment or '')[:200],
        ], fill, WRAP)
        # Жирный Макс
        if in_mx:
            c = ws1.cell(row=ri, column=13)
            c.font = Font(bold=True, color="7F6000"); c.fill = MAXIM_F

    # ══════════════════════════════════════════════════════════════════════
    # Лист 2: «Родственники»
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Родственники")
    cols2 = [
        ("Исходное ФИО",              26),
        ("Найденный человек",         26),
        ("Родственник ФИО",           28),
        ("Дата рождения",             14),
        ("Прежняя/другая фамилия",    22),
        ("Тип родства / причина",     22),
        ("Адрес родственника",        38),
        ("Телефоны родственника",     30),
        ("Проверка номера",           22),
        ("ВКонтакте",                 22),
        ("Одноклассники",             22),
        ("Другие соцсети",            22),
        ("Есть в Максе",              10),
        ("Источник Макса",            25),
        ("Доказательство родства",    35),
        ("Уверенность",               12),
        ("Комментарий",               28),
    ]
    _write_header(ws2, cols2, H2, HFNT, WRAP)

    for ri, r in enumerate(relatives, 2):
        if r.confidence == "высокая":
            fill2 = REL_HIGH
        elif r.in_maxim:
            fill2 = REL_MX
        else:
            fill2 = REL_BG
        _write_row(ws2, ri, [
            r.source_fio, r.main_fio, r.fio,
            r.dob, r.alt_fio, r.relation,
            r.address,
            '; '.join(r.phones),
            r.phone_note,
            r.vk, r.ok, r.other_social,
            "✅ Да" if r.in_maxim else "Нет",
            r.maxim_source,
            r.evidence, r.confidence, r.comment,
        ], fill2, WRAP)
        if r.in_maxim:
            c = ws2.cell(row=ri, column=13)
            c.font = Font(bold=True, color="7F6000"); c.fill = REL_MX

    # ══════════════════════════════════════════════════════════════════════
    # Лист 3: «Проверка номеров»
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Проверка номеров")
    cols3 = [
        ("Кому принадлежит",     28),
        ("Тип",                  14),
        ("Номер (норм.)",        18),
        ("Оператор/тип",         16),
        ("В Sauron",             10),
        ("Есть в Максе",         10),
        ("Соцсети",              22),
        ("Свежие источники",     12),
        ("Статус",               16),
        ("Причина отбора/отброса",28),
    ]
    _write_header(ws3, cols3, H3, HFNT, WRAP)

    for ri, pc in enumerate(phones, 2):
        is_mobile = _is_mobile(pc.phone_norm)
        op = "мобильный" if is_mobile else "городской"
        fill3 = PH_ACT if pc.status == "действующий" else PH_DIS
        _write_row(ws3, ri, [
            pc.owner_fio, pc.owner_type, pc.phone_norm,
            op,
            "Да" if pc.in_sauron else "Нет",
            "✅ Да" if pc.in_maxim else "Нет",
            pc.social_links,
            "✅" if pc.fresh_src else "—",
            pc.status,
            pc.reject_reason or ("Топ по скорингу" if pc.status=="действующий" else ""),
        ], fill3, WRAP)

    # ══════════════════════════════════════════════════════════════════════
    # Лист 4: «Ошибки и лимиты»
    # ══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Ошибки и лимиты")
    cols4 = [
        ("Строка №",    8),
        ("ФИО",        28),
        ("Тип",        18),
        ("Детали",     60),
    ]
    _write_header(ws4, cols4, H4, HFNT, WRAP)

    for ri, e in enumerate(errors, 2):
        _write_row(ws4, ri, [
            e.row_num, e.source_fio, e.error_type, e.detail,
        ], ERR_F, WRAP)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═════════════════════════════════════════════════════════════════════════════
# FINAL_MERGED.xlsx — основной формат результата (один лист, 1 строка = 1 родственник)
# ═════════════════════════════════════════════════════════════════════════════

FINAL_MERGED_HEADERS = [
    "ФИО погибшего",
    "Дата рождения погибшего",
    "Дата смерти погибшего",
    "ФИО родственника",
    "Дата рождения родственника",
    "Телефон",
    "phone_norm",
    "СНИЛС",
    "MAX",
    "Соц. сети",
    "Ошибка",
]
# Колонки, которые Excel должен хранить как текст (чтобы не портить длинные числа)
_FM_TEXT_COLS = {6, 7, 8, 9}  # Телефон, phone_norm, СНИЛС, MAX


def _fm_social_links(r: RelativeRecord) -> str:
    """Собирает ссылки VK/OK/Telegram/Instagram, каждая с новой строки."""
    links: list[str] = []
    for raw in (r.vk_profile_url, r.vk, r.ok, r.other_social):
        if not raw:
            continue
        for part in re.split(r'[;\n\s]+', str(raw)):
            part = part.strip()
            if part and part not in links:
                links.append(part)
    return '\n'.join(links)


def build_final_merged_xlsx(
    persons:   list[PersonRecord],
    relatives: list[RelativeRecord],
    phones:    list[PhoneCheck],
) -> Optional[bytes]:
    """
    Единый файл FINAL_MERGED.xlsx, лист Sheet1.
    Одна строка = один родственник найденного (погибшего) человека.
    Дополнительно: строки для не найденных / ошибочных людей (с заполненной «Ошибка»).
    """
    if not _HAS_OPENPYXL:
        return None

    # Индекс: строка/ФИО исходного человека → PersonRecord (для даты рожд./смерти)
    # Приоритет — точный номер строки, fallback — ФИО (на случай старых записей).
    person_by_row: dict[int, PersonRecord] = {p.row_num: p for p in persons if p.row_num}
    person_by_src: dict[str, PersonRecord] = {}
    for p in persons:
        person_by_src.setdefault(p.source_fio.strip().lower(), p)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    head_fill = PatternFill("solid", fgColor="1F4E79")
    head_font = Font(bold=True, color="FFFFFF")
    wrap_top  = Alignment(wrap_text=True, vertical="top")

    for ci, title in enumerate(FINAL_MERGED_HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=title)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    widths = [30, 18, 18, 30, 18, 34, 28, 18, 26, 36, 24]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    def _emit(row_idx: int, values: list[str]):
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=ci, value=val if val else None)
            c.alignment = wrap_top
            if ci in _FM_TEXT_COLS:
                c.number_format = '@'   # хранить как текст

    ri = 2

    # ── Одна строка = один родственник найденного (погибшего) человека ─────
    for r in relatives:
        p = person_by_row.get(r.source_row) or person_by_src.get(r.source_fio.strip().lower())

        deceased_fio = (r.main_fio or r.source_fio or "").strip()
        deceased_dob = p.dob if p else ""
        deceased_dod = p.dod if p else ""

        # Телефон с операторами: "79140367459 (МТС); 79512941691 (Tele2)"
        phone_disp_parts: list[str] = []
        for n in r.phones:
            op = _phone_operator(n)
            phone_disp_parts.append(f"{n} ({op})" if op else n)
        phone_disp = '; '.join(phone_disp_parts)
        phone_norm = '; '.join(r.phones)

        # MAX: "Да - 79148605668" если человек в Максе (флаг — на уровне родственника).
        # in_maxim не привязан к конкретному номеру, поэтому берём первый ликвидный.
        if r.in_maxim and r.phones:
            max_cell = "Да - " + r.phones[0]
        elif r.in_maxim:
            max_cell = "Да"
        else:
            max_cell = "Нет"

        _emit(ri, [
            deceased_fio,
            deceased_dob,
            deceased_dod,
            r.fio,
            r.dob,
            phone_disp,
            phone_norm,
            r.snils,
            max_cell,
            _fm_social_links(r),
            "",   # ошибок по строке родственника нет
        ])
        ri += 1

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═════════════════════════════════════════════════════════════════════════════
# CSV / ZIP отчёт (вспомогательный экспорт)
# ═════════════════════════════════════════════════════════════════════════════

def _make_csv(headers: list[str], rows: list[list]) -> bytes:
    """Один CSV-файл: UTF-8 BOM + точка-с-запятой (корректно открывается в Excel RU)."""
    out = io.StringIO()
    w   = csv.writer(out, delimiter=';', quoting=csv.QUOTE_ALL)
    w.writerow(headers)
    for row in rows:
        w.writerow([str(v) if v is not None else '' for v in row])
    return out.getvalue().encode('utf-8-sig')


def _csv_people(persons: list[PersonRecord]) -> bytes:
    headers = [
        "Строка №", "Исходное ФИО", "Дата рождения", "Дата смерти",
        "Найдено", "Найденное ФИО",
        "Адреса", "Лучшие телефоны", "Email",
        "ВКонтакте", "Одноклассники", "Другие соцсети",
        "Есть в Максе", "Источник Макса",
        "Источники", "Родственников найдено", "Ошибка / Комментарий",
    ]
    rows = []
    for p in persons:
        rows.append([
            p.row_num, p.source_fio, p.dob, p.dod,
            "Да" if p.found else "Нет", p.found_fio,
            p.addresses, p.phones, p.emails,
            p.vk, p.ok, p.other_social,
            "Да" if p.in_maxim else "Нет", p.maxim_source,
            p.sources[:300] if p.sources else '',
            p.rel_count,
            p.error or p.comment,
        ])
    return _make_csv(headers, rows)


def _csv_relatives(relatives: list[RelativeRecord]) -> bytes:
    headers = [
        "Исходное ФИО", "Найденный человек",
        "Родственник ФИО", "Дата рождения родственника",
        "Варианты ФИО / алиасы", "Тип родства / причина",
        "Адрес родственника", "Действующие телефоны", "Примечание по телефонам",
        "ВКонтакте (Sauron)", "Одноклассники", "Другие соцсети",
        "Есть в Максе", "Источник Макса",
        "Доказательства из соцсетей", "Общие признаки",
        "Скоринг родства", "Источник доказательства",
        "Все доказательства родства", "Уверенность",
        # VK API колонки
        "VK профиль", "VK имя", "VK девичья/старая фамилия",
        "VK родственники", "VK город", "Доказательства VK",
        "Комментарий",
    ]
    rows = []
    for r in relatives:
        rows.append([
            r.source_fio, r.main_fio,
            r.fio, r.dob,
            r.variants_fio or r.alt_fio, r.relation,
            r.address, '; '.join(r.phones), r.phone_note,
            r.vk, r.ok, r.other_social,
            "Да" if r.in_maxim else "Нет", r.maxim_source,
            r.social_evidence, r.common_signs,
            r.score, r.evidence_source,
            r.evidence, r.confidence,
            r.vk_profile_url, r.vk_full_name, r.vk_maiden_name,
            r.vk_relatives_str, r.vk_city, r.vk_evidence,
            r.comment,
        ])
    return _make_csv(headers, rows)


def _csv_phones(phones: list[PhoneCheck]) -> bytes:
    headers = [
        "Кому принадлежит", "Тип (основной / родственник)",
        "Номер (нормализованный)", "Тип номера",
        "Найден в Sauron", "Есть в Максе",
        "Соцсети", "Свежие источники (2024-2026)",
        "Статус", "Причина отбора / отброса", "Подтверждено источниками",
    ]
    rows = []
    for pc in phones:
        is_mob = _is_mobile(pc.phone_norm)
        confirmed = []
        if pc.in_sauron:    confirmed.append('Sauron')
        if pc.in_maxim:     confirmed.append('Макс/taxsee')
        if pc.fresh_src:    confirmed.append('свежий источник 2024+')
        if pc.social_links: confirmed.append('соцсеть')
        rows.append([
            pc.owner_fio, pc.owner_type,
            pc.phone_norm,
            "мобильный" if is_mob else "городской",
            "Да" if pc.in_sauron else "Нет",
            "Да" if pc.in_maxim else "Нет",
            pc.social_links,
            "Да" if pc.fresh_src else "Нет",
            pc.status,
            pc.reject_reason or ("Топ по скорингу" if pc.status == "действующий" else ""),
            '; '.join(confirmed),
        ])
    return _make_csv(headers, rows)


def _csv_errors(errors: list[ErrorRecord]) -> bytes:
    headers = ["Строка №", "ФИО / Файл", "Тип ошибки", "Детали"]
    rows = [[e.row_num, e.source_fio, e.error_type, e.detail] for e in errors]
    return _make_csv(headers, rows)


def build_zip_report(
    persons:   list[PersonRecord],
    relatives: list[RelativeRecord],
    phones:    list[PhoneCheck],
    errors:    list[ErrorRecord],
    base_name: str = "sauron",
) -> bytes:
    """
    Собирает ZIP-архив с 4 CSV-файлами:
      people_summary.csv   — Итог по людям
      relatives.csv        — Родственники
      phone_checks.csv     — Проверка номеров
      errors_limits.csv    — Ошибки и лимиты

    Кодировка: UTF-8 BOM, разделитель «;» (корректно открывается в Excel RU).
    """
    import zipfile

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{base_name}_people_summary.csv",  _csv_people(persons))
        zf.writestr(f"{base_name}_relatives.csv",       _csv_relatives(relatives))
        zf.writestr(f"{base_name}_phone_checks.csv",    _csv_phones(phones))
        zf.writestr(f"{base_name}_errors_limits.csv",   _csv_errors(errors))
    return buf.getvalue()


def build_csv_report(
    persons:   list[PersonRecord],
    relatives: list[RelativeRecord],
) -> bytes:
    """Устаревший единый CSV (для обратной совместимости). Предпочитай build_zip_report."""
    return _make_csv(
        [
            "Исходное ФИО", "Строка №", "Дата рождения", "Дата смерти", "Найдено",
            "Найденное ФИО", "Адреса", "Телефоны", "Email", "ВК", "ОК",
            "В Максе", "Источник Макса", "Источники", "Родственников", "Ошибка",
            "Родств. ФИО", "Родств. ДР", "Прежняя фамилия", "Тип родства",
            "Адрес родств.", "Тел. родств.", "Родств. ВК", "Родств. ОК",
            "Родств. в Максе", "Доказательство", "Уверенность",
        ],
        _flat_rows(persons, relatives),
    )


def _flat_rows(persons, relatives):
    rel_by = {}
    for r in relatives:
        rel_by.setdefault(r.source_fio.lower(), []).append(r)
    rows = []
    for p in persons:
        rels = rel_by.get(p.source_fio.lower(), [])
        base = [
            p.source_fio, p.row_num, p.dob, p.dod,
            "Да" if p.found else "Нет", p.found_fio,
            p.addresses, p.phones, p.emails, p.vk, p.ok,
            "Да" if p.in_maxim else "Нет", p.maxim_source,
            p.sources, p.rel_count, p.error or p.comment,
        ]
        if rels:
            for r in rels:
                rows.append(base + [
                    r.fio, r.dob, r.alt_fio, r.relation,
                    r.address, '; '.join(r.phones),
                    r.vk, r.ok,
                    "Да" if r.in_maxim else "Нет",
                    r.evidence, r.confidence,
                ])
        else:
            rows.append(base + [''] * 11)
    return rows


# ═════════════════════════════════════════════════════════════════════════════
# Превью до начала поиска
# ═════════════════════════════════════════════════════════════════════════════

def build_preview(records: list[InputRecord], filename: str, max_fio: int = MAX_FIO) -> str:
    total  = len(records)
    will   = min(total, max_fio)
    skip   = total - will
    lines  = [f"📄 *Файл:* `{filename}`\n"]
    lines.append(f"👤 Найдено ФИО: *{total}*")
    has_dob = sum(1 for r in records if r.dob)
    has_dod = sum(1 for r in records if r.dod)
    if has_dob:
        lines.append(f"📅 С датой рождения: {has_dob}")
    if has_dod:
        lines.append(f"✝️ Дата смерти указана: {has_dod}")
    lines.append("")
    for r in records[:8]:
        dob_str = f"  _{r.dob}_" if r.dob else ""
        lines.append(f"  • {r.fio}{dob_str}")
    if total > 8:
        lines.append(f"  _…ещё {total-8}_")
    lines.append("\n━━━━━━━━━━━━━━━━━━━━")
    lines.append(f"🔍 Будет проверено: *{will}* ФИО")
    if skip:
        lines.append(f"⚠️ По лимиту ({max_fio}) пропущено: *{skip}*")
    return '\n'.join(lines)


def supported_formats() -> str:
    fmts = ["txt", "csv"]
    if _HAS_OPENPYXL:
        fmts.append("xlsx")
    if _HAS_XLRD:
        fmts.append("xls")
    if _HAS_DOCX:
        fmts.append("docx")
    if _HAS_PDF:
        fmts.append("pdf")
    return ", ".join(fmts)
